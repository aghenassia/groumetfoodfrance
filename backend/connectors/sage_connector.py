"""
Connecteur Sage 100.

Mode ODBC : connexion directe au SQL Server Sage (production)
Mode Excel : import depuis fichiers d'extraction (dev/seed)
"""
import re
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.dialects.postgresql import insert as pg_insert

from models.client import Client
from models.sales_line import SalesLine
from models.phone_index import PhoneIndex
from models.sync_log import SyncLog
from connectors.phone_normalizer import normalize_phone, extract_all_phones


def clean_sage_string(val) -> str | None:
    """Nettoie une valeur issue de Sage/Excel (enlève ="...", espaces, etc.)."""
    if val is None:
        return None
    s = str(val).strip()
    s = re.sub(r'^="?|"$', "", s)
    s = s.strip()
    return s if s and s != "None" else None


def clean_sage_number(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


async def import_clients_from_excel(db: AsyncSession, filepath: str) -> dict:
    """Importe les clients depuis un fichier Excel extrait de Sage."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    created, updated, errors = 0, 0, 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        try:
            sage_id = clean_sage_string(data.get("Numéro"))
            if not sage_id:
                continue

            name = clean_sage_string(data.get("Intitulé")) or "Inconnu"
            phone_raw = clean_sage_string(data.get("Téléphone"))
            country = clean_sage_string(data.get("Pays")) or "France"
            email_raw = clean_sage_string(data.get("E-mail"))
            # Certains emails Sage contiennent plusieurs adresses séparées par ";"
            email = email_raw.split(";")[0].strip() if email_raw else None

            sage_created = None
            raw_date = data.get("Date de création")
            if raw_date and isinstance(raw_date, datetime):
                sage_created = raw_date.replace(tzinfo=timezone.utc)

            client_values = {
                "sage_id": sage_id,
                "name": name,
                "short_name": clean_sage_string(data.get("Abrégé")),
                "quality": clean_sage_string(data.get("Qualité")),
                "contact_name": clean_sage_string(data.get("Contact")),
                "address": clean_sage_string(data.get("Adresse")),
                "address_complement": clean_sage_string(data.get("Complément adresse")),
                "postal_code": clean_sage_string(data.get("Code postal")),
                "city": clean_sage_string(data.get("Ville")),
                "region": clean_sage_string(data.get("Région")),
                "country": country,
                "phone": phone_raw,
                "phone_e164": normalize_phone(phone_raw, _country_to_region(country)),
                "fax": clean_sage_string(data.get("Télécopie")),
                "email": email,
                "website": clean_sage_string(data.get("Site")),
                "siret": clean_sage_string(data.get("N° Siret")),
                "vat_number": clean_sage_string(data.get("N° identifiant")),
                "naf_code": clean_sage_string(data.get("Code NAF")),
                "sales_rep": clean_sage_string(data.get("Représentant")),
                "tariff_category": clean_sage_string(data.get("Catégorie tarifaire")),
                "accounting_category": clean_sage_string(data.get("Catégorie comptable")),
                "is_prospect": bool(data.get("Prospect")),
                "is_dormant": bool(data.get("Mise en sommeil")),
                "sage_created_at": sage_created,
                "synced_at": datetime.now(timezone.utc),
            }

            # Upsert client — merge : on ne met à jour que les champs
            # qui ont une valeur dans Sage (on n'écrase jamais un champ
            # CRM non-vide avec un null Sage)
            update_set = {}
            for k, v in client_values.items():
                if k in ("sage_id", "synced_at"):
                    continue
                if v is not None:
                    update_set[k] = v
            update_set["synced_at"] = client_values["synced_at"]

            stmt = pg_insert(Client).values(**client_values)
            stmt = stmt.on_conflict_do_update(
                index_elements=["sage_id"],
                set_=update_set,
            )
            result = await db.execute(stmt)

            # Récupérer le client pour son ID
            q = await db.execute(select(Client).where(Client.sage_id == sage_id))
            client = q.scalar_one()

            # Mettre à jour le phone_index
            phone_entries = extract_all_phones({
                "phone": phone_raw,
                "fax": clean_sage_string(data.get("Télécopie")),
                "country": country,
            })
            for phone_entry in phone_entries:
                phone_stmt = pg_insert(PhoneIndex).values(
                    phone_e164=phone_entry["phone_e164"],
                    client_id=client.id,
                    source=phone_entry["source"],
                    raw_phone=phone_entry["raw_phone"],
                    label=phone_entry["label"],
                )
                phone_stmt = phone_stmt.on_conflict_do_nothing(
                    constraint="uq_phone_client"
                )
                await db.execute(phone_stmt)

            created += 1

        except Exception as e:
            await db.rollback()
            errors += 1
            print(f"Erreur import client: {e}")

    if created > 0:
        await db.commit()

    wb.close()
    return {"created": created, "updated": updated, "errors": errors}


async def import_sales_from_excel(db: AsyncSession, filepath: str) -> dict:
    """Importe les lignes de ventes depuis un fichier Excel extrait de Sage."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    created, errors = 0, 0

    # Pré-charger le mapping sage_id → client.id
    q = await db.execute(select(Client.sage_id, Client.id))
    client_map = {row[0]: row[1] for row in q.all()}

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        try:
            piece_id = clean_sage_string(data.get("N° pièce"))
            client_sage_id = clean_sage_string(data.get("N° tiers entête"))
            if not piece_id or not client_sage_id:
                continue

            raw_date = data.get("Date")
            if isinstance(raw_date, datetime):
                line_date = raw_date.date()
            else:
                continue

            article_ref = clean_sage_string(data.get("Référence article"))

            values = {
                "sage_piece_id": piece_id,
                "sage_doc_type": 6,
                "client_sage_id": client_sage_id,
                "client_id": client_map.get(client_sage_id),
                "client_name": clean_sage_string(data.get("Intitulé tiers ligne")),
                "date": line_date,
                "article_ref": article_ref,
                "designation": clean_sage_string(data.get("Désignation")),
                "quantity": clean_sage_number(data.get("Quantité")),
                "unit_price": clean_sage_number(data.get("P.U. HT")),
                "net_weight": clean_sage_number(data.get("Poids net global")),
                "discount": clean_sage_number(data.get("Remise")),
                "net_unit_price": clean_sage_number(data.get("P.U. net")),
                "cost_price": clean_sage_number(data.get("Prix de revient unitaire")),
                "amount_ht": clean_sage_number(data.get("Montant HT")) or 0,
                "sales_rep": clean_sage_string(data.get("Collaborateur")),
                "margin_value": clean_sage_number(data.get("Marge en valeur")),
                "margin_percent": clean_sage_number(data.get("Marge en %")),
                "synced_at": datetime.now(timezone.utc),
            }

            stmt = pg_insert(SalesLine).values(**values)
            stmt = stmt.on_conflict_do_update(
                constraint="uq_sales_line",
                set_={k: v for k, v in values.items()
                      if k not in ("sage_piece_id", "article_ref", "client_sage_id", "date")},
            )
            await db.execute(stmt)
            created += 1

            if created % 100 == 0:
                await db.commit()

        except Exception as e:
            await db.rollback()
            errors += 1
            if errors <= 3:
                print(f"Erreur import vente: {e}")

    await db.commit()

    wb.close()
    return {"created": created, "errors": errors}


def _country_to_region(country: str) -> str:
    mapping = {
        "france": "FR", "allemagne": "DE", "espagne": "ES",
        "luxembourg": "LU", "pays-bas": "NL", "serbie": "RS",
    }
    return mapping.get(country.lower().strip(), "FR")
